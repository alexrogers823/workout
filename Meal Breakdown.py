from openpyxl.cell import get_column_letter
from openpyxl.styles import Font, Style, Color, PatternFill
import openpyxl as px
import shift, time, subprocess

def set_workbook():
    food = wb.get_sheet_by_name('Meal Index')
    food['A1'] = 'Weight:'
    food['D1'] = 'Daily Group Goals:'
    for i in range(len(titles)):
        food[get_column_letter(i+1)+'3'] = titles[i]
        food[get_column_letter(i+1)+'3'].style = Style(font=Font(bold=True, size=12))
    food.column_dimensions['B'].width = 18
    food.column_dimensions['C'].width = 18
    food.column_dimensions['D'].width = 33
    food.freeze_panes = 'A4'
    return food

def weight():
    global wt
    print('Assuming a weight of %s... [enter correct weight if incorrect]' % (wt))
    if input().isdigit():
        wt = int(input())
        food['B1'] = wt
    add_meal()
    return

def change_weight(): #Obsolete
    global wt
    print('Enter current weight')
    wt = int(input())
    food['B1'] = wt
    return

def add_meal():
    global x, refs_this_session
    x += 1
    if len(refs_this_session) > 0:
        print('Add next one: (? to bring up order)')
    if len(refs_this_session) == 0: #or input() == '?':
        print('Add the following, separating each by a comma:')
        for i in range(len(titles)-2):
            print(titles[i], end=', ')
        print()
    while True:
        try:
            meal = input().title().split(', ') #separate meal so each part goes in the right place on excel sheet
            dup = check_duplicates(meal[3])
            if dup == True:
                raise Exception
        except IndexError:
            print('Add ALL of the following:')
            for i in range(len(titles)-2):
                print(titles[i], end=', ')
            print()
        except Exception:
            print('Try another title')
            time.sleep(1)
        else:
            break
    mealType = meal[0]
    r = categorize_meal_type(mealType)
    for j in range(len(titles)-2):
        food[get_column_letter(j+1)+str(r)] = meal[j]
    j += 1
    print('Is this a favorite?')
    if input().lower().startswith('y'):
        food[get_column_letter(j+1)+str(r)] = 'x'
    j += 1
    print('Where is the recipe from?')
    if len(refs_this_session) > 0:
        last = refs_this_session[len(refs_this_session)-1]
        print('(Last input: {})'.format(last))
    rec = input()
    if rec.isdigit():
        #copy the one before
        food[get_column_letter(j+1)+str(r)] = last[:(last.rindex('p.')+3)]+rec
    else:
        food[get_column_letter(j+1)+str(r)] = rec
    refs_this_session.append(food[get_column_letter(j+1)+str(r)].value)
    more()

def more():
    print('Others to add?')
    if input().lower().startswith('n'):
        return
    else:
        add_meal()

def check_duplicates(name, dup=False):
    for i in range(4, x+1):
        if name == food['D'+str(i)].value:
            print('{} already exists'.format(name))
            dup = True
            break
        else:
            continue
    return dup

def categorize_meal_type(search):
    cate = ['Breakfast', 'Lunch', 'Dinner', 'Snack', 'Dessert']
    for i in range(4, x+1):
        if search == food['A'+str(i)].value:
            break
        else:
            continue
    r = i
    shift.Shift(food, food.max_column, x, r).down()
    return r

def set_goals():
    cal = 15
    protein = 1
    carbs = 2
    fat = 0.4
    gNum = [cal, protein, carbs, fat]
    food['F1'] = '{} cal'.format(str(wt*gNum[0])) #Calories
    for i in range(7, 10): #Everything else
        food[get_column_letter(i)+'1'] = '{} g'.format(str(wt*gNum[i-7]))
    return
    

def save_and_update():
    wb.save('Meal Breakdown.xlsx')
    opener = 'open'
    print('Opening...')
    time.sleep(2)
    subprocess.call([opener, ('Meal Breakdown.xlsx')])
    
#Main
refs_this_session = []
titles = ['Meal Type', 'Primary Food Type', 'Second Food Type', 'Name', 'Unit', 'Calories', 'Protein (g)', 'Carbs (g)', 'Fat (g)', 'Favorites', 'Reference']
wb = px.load_workbook('Meal Breakdown.xlsx')
food = set_workbook()
wt = food['B1'].value
x = food.max_row
weight()
set_goals()
save_and_update()
