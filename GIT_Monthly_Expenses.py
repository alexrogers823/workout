#from apscheduler.schedulers.background import BackgroundScheduler
from openpyxl.styles import Font, Style, Color, PatternFill
from openpyxl.cell import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference, Series
from openpyxl.chart.layout import Layout , ManualLayout
from openpyxl.chart.label import DataLabel, DataLabelList
import openpyxl as px
#import save_and_update as saving
import sys, os, subprocess, time, datetime, random, webbrowser, shift

# This def will pull up your banks with webbrowser and ask you to login
# It also tells you how many days until your next payment is due
def reviewMaterials():
    banks = websites()
    print('Login into your accounts')
    # dNum is the todays date expressed as a number. A global variable
    if dNum <= 12:
        WF = 23 - dNum
        Dis = 12 - dNum
    elif dNum > 12 and dNum <= 23:
        WF = 23 - dNum
        Dis = '20+'
    else:
        WF = '20+'
        Dis = '10+'
    print('REMINDER: You have %s days to pay off your WF balance,\n and %s days to pay off your Discover balance' % (WF, Dis))
    print()
    print('[Disabled webbrowser for testing]')
    time.sleep(1)
    #for i in banks:
#        webbrowser.open(banks[i])
#    webbrowser.open('https://www.discover.com')
#    webbrowser.open('https://www.wellsfargo.com')
#    webbrowser.open('https://www.capitalone.com')
    # Will bring up last expenses you input in excel, and will tell you last
    # two expenses in the shell
    print('Do you need to see the excel sheet as well? [Y or N]')
    if x > 2:
        print('(The last two inputs were {} for ${}, and {} for ${})'.format(month['A'+str(x)].value, month['D'+str(x)].value, month['A'+str(x-1)].value, month['D'+str(x-1)].value))
    if input().lower().startswith('y'):
        opener = 'open'
        print('Reviewing...')
        time.sleep(2)
        subprocess.call([opener, (curYear + '.xlsx')])

def websites(): #Manages websites used to review payments
    names = []
    sites = []
    w = 2
    while Sum[get_column_letter(mNum+14)+str(w)].value != None:
        names.append(Sum[get_column_letter(mNum+14)+str(w)].value)
        w += 1
    if len(names) == 0: #or add.lower().startswith('ad'):
        print('Bank website setup:')
        print()
        print('Please enter the FULL NAME of the banks you use, separated by a comma (,)')
        names = input().split(',')
    for i in names:
        sites.append('https://www.{}.com'.format(names[i].lower().replace(' ', '')))

    return sites

# Checks if this is the first input this year, and makes new workbook if so
def set_new_year():
    ny = px.load_workbook('blank.xlsx')
    ny.create_sheet(index = 0, title = 'Summary')
    ny.get_sheet_by_name('Summary')['J1'] = 'Average'
    ny.create_sheet(index = 1, title = 'Category Words')
    ny.save(str(int(year)+1) +' Monthly Expenses.xlsx')
    return

# Checks if this is first input of the month, and makes new sheet if so 
def setWorkbook():
    if Sum[get_column_letter(mNum+9) + '1'].value == 'Average':
        if mNum == 12: #mNum is current month expressed as a number. See Main
            set_new_year()
        wb.create_sheet(index = mNum, title = TM[mNum])
        shift.Shift(Sum, Sum.max_column, Sum.max_row, mNum+12).right()
    currentMonth = wb.get_sheet_by_name(TM[mNum])
    return currentMonth

# Resets look of summary sheet 
def setSummarySheet():
    styleYear = Style(font=Font(name='Calibri', bold=True, color='009999'))
    mStyle = Style(font=Font(name='Calibri', color='FFFFFF'))
    mFill = PatternFill(start_color='2E2EB8', fill_type='solid')
    cate = ['Rent', 'Utilities', 'Phone', 'Apparel','Supplies', 'Technology', 'Services', 'Health/Gym', 'Haircut', 'Groceries', 'Other Food', 'Gas & Parking', 'Insurance','Bank & Credit Card', 'Student Loans', 'Vehicle Payments', 'Entertainment', 'Subscriptions', 'Travel', 'Public Transportation', 'Special/Seasonal', 'Other', '', 'Total Expenses']
    for i in range(0, mNum+1): #Resets font of previous months
        Sum[get_column_letter(i+9) + '1'] = TM[i]
        Sum[get_column_letter(i+9) + '1'].style = GA

    Sum[get_column_letter(mNum+10) + '1'] = 'Average'
    Sum[get_column_letter(mNum+11) + '1'] = '3-Mo. Avg'
    Sum[get_column_letter(mNum+12) + '1'] = 'Month Goal'
    for a in range(10, 12): #Resets font of 'Average', '3-Mo. Avg', and 'Month Goal'
        Sum[get_column_letter(mNum+a) + '1'].style = Style(font=Font(name='Georgia', bold=True, color='AB800A'))

    Sum[get_column_letter(mNum+12)+'1'].style = Style(font=Font(name='Impact', bold=True, color='2E2EB8'))
    
    # Gives different font color for current month
    Sum[get_column_letter(mNum + 9) +'1'].style = Style(font=Font(name='Georgia', bold=True, color='2E2EB8'))
    
    for j in range(len(cate)): #Resets look of categories
        Sum['H' + str(j+2)] = cate[j] + ':'
        Sum['H' + str(j+2)].style = GA
        Sum['I' + str(j+2)].style = styleYear
        Sum[get_column_letter(mNum+10) + str(j+2)].style = Style(font=Font(name='Calibri', bold=True, color='AB800A'))
        Sum[get_column_letter(mNum+11) + str(j+2)].style = Style(font=Font(name='Calibri', bold=True, color='AB800A'))
        Sum[get_column_letter(mNum + 9) + str(j+2)].style = mStyle
        Sum[get_column_letter(mNum + 9) + str(j+2)].fill = mFill
        Sum[get_column_letter(mNum + 9) + str(len(cate))].fill = PatternFill(fill_type =None)

    for e in range(len(TM)): #Separates space between 'Other' and 'Total'
        Sum[get_column_letter(e+8) +str(len(cate))] = ''

    cword['A1'] = 'Keyword'
    cword['A1'].style = GA
    cword['B1'] = 'Category'
    cword['B1'].style = GA
        
    return cate
       
def formatSheet(month):
    month.column_dimensions['A'].width = 23
    month.column_dimensions['B'].width = 15
    month.column_dimensions['C'].width = 23
    month.column_dimensions['E'].width = 20
    Sum.column_dimensions['H'].width = 23

    if x > 29: #Freezes panes when month inputs are longer than sheet
        month.freeze_panes = 'A2'

    styleGA = Style(font=Font(name='Georgia', bold=True))

    return styleGA

def setMonth():
    labels = ['Expense', 'Category', 'Notes', 'Price', 'Cumulative Price']
    for i in range(len(labels)):
        month['{}1'.format(chr(i+65))] = labels[i]
        month['{}1'.format(chr(i+65))].style = GA
        
 #   month['A1'] = 'Expense'
#    month['A1'].style = GA
#    month['B1'] = 'Category'
#    month['B1'].style = GA
#    month['C1'] = 'Notes'
#    month['C1'].style = GA
#    month['D1'] = 'Price'
#    month['D1'].style = GA
#    month['E1'] = 'Cumulative Price'
#    month['E1'].style = GA
    month['G1'] = '*Most recent expenses in bold text (Scroll Down)'
    return

def revertStyle(): #Resets previous expenses to normal font
    standard = Style(font=Font(name='Calibri', bold=False))
    for i in range(2,(x+1)):
        for j in range(0,5):
            month[get_column_letter(j+1) + str(i)].style = standard

    for a in range(len(categories)): #Resets previous month summary to normal font
        for b in range(1, mNum):
            Sum[get_column_letter(b+9) + str(a+2)].style = standard

def checkCate(): #Checks if name of input is already in a category
    d = cword.max_row
    theCate = ''
    for i in range(2, d+1):
        if cword['A'+str(i)].value == keyword:
            theCate = cword['B'+str(i)].value

    return theCate
            

def addToCate(): #Adds new keyword to category upon request
    d = cword.max_row
    d += 1
    cword['A'+str(d)] = keyword
    print('Which category to assign?')
    print()
    for i in range(len(categories)-3):
        print(i, categories[i])
    assign = int(input())
    cword['B'+str(d)] = categories[assign]
    
# Major part of code. This is where users input each expense
def addExpense():
    global x, keyword
    x += 1
    current = Style(font=Font(name='Cooper Black', bold=False))

    # Generates random examples based on your previous expenses
    Ex1 = random.randint(0, 4)
    Ex2 = random.randint(5, 9)
    Ex3 = random.randint(10, len(categories) - 3)
    names = []
    for i in range(3):
        names.append(month['A'+str(random.randint(1, x-1))].value)

    if x < 5:
        print('What\'s the name of your expense? (Ex: eggs, clothes, vodka)')
    else:
        print('What\'s the name of your expense? (Ex: {0[0]}, {0[1]}, {0[2]})'.format(names))
    exp = input()
    if exp.lower().startswith('correct'): #Short exits if small correction in code was made
        print('Correction initiated')
        shortExit()
        sys.exit()
    if exp.lower() == 'no': #When user doesn't have another expense but accidently said they did
        shortExit()
        sys.exit()

    month['A' + str(x)] = exp
    keyword = exp
    val = checkCate() #Checks if the name of the expense is already in category
    print()
    if val != '':
        month['B'+str(x)] = val.title()
    else:
        print('Under which category? (Ex: %s, %s, %s)' % (categories[Ex1], categories[Ex2], categories[Ex3]))
        print('(* at end to assign)')
        cat = input()
        if cat.endswith('*'):
            addToCate()
        cat = cat.strip('*')
        print()
        if cat == 'food':
            print('Was it food from the grocery store? [Y or N]')
            if input().lower().startswith('y'):
                month['B' +str(x)] = 'Groceries'
            else:
                month['B' +str(x)] = 'Res'
            print()
        elif cat == 'cvs':
            print('Was it health related? (if for supplies then N) [Y or N]')
            if input().lower().startswith('y'):
                month['B' +str(x)] = 'Health'
            else:
                month['B' +str(x)] = 'Supplies'
            print()
        else:
            month['B' + str(x)] = cat.title()
    print('Additional notes?')
    if dNum < 10:
        print('(if from previous month, put \'qq\' here)') #Will calculate this expense with previous month
    month['C' + str(x)] = input()
    print()
    while True:
        print('Total cost? (Enter as x.xx)')
        cost = input()
        try: 
            month['D' + str(x)] = float(cost)
        except ValueError: #Prevents crash of code if input isn't a number
            print('Be careful! You have to enter a number here')
        else:
            break

    if ['B'+str(x)] == 'Other Food' and float(cost) > float(25):
        print('Did you pay for others\' food?')
        if input().lower().startswith('y'):
            month['D'+str(x)].fill = PatternFill(start_color='6DC066', fill_type='solid')
    print()

    for i in range(5): #Puts new expense in bold text
        month[get_column_letter(i+1) + str(x)].style = current
            

# Recalculates all category totals in month and year after new expenses
def categoryTotal():
    ml = len(categories)
    for g in range(1, mNum+1):
        mez = wb.get_sheet_by_name(TM[g])
        mz = g+1
        z = mez.max_row
        for i in range(ml-2):
            Sum[get_column_letter(mz+8) + str(i+2)] = 0 #Resets numbers in summary to zero
 #           Sum['I' +str(i+2)] = '=SUM(J' +str(i+2)+ ':'+get_column_letter(mNum+9) +str(i+2)+ ')'
            Sum['I'+str(i+2)] = '=SUM(J{0}:{1}{0})'.format(str(i+2), get_column_letter(mNum+9))
            y = i+1
            if y == 1: #Rent
                ct = ['Rent','Lease',]
            if y == 2: #Utilities
                ct = ['Util','Utility','Utilities','Water Bill','Electricity', 'Water', 'Bill']
            if y == 3: #Phone
                ct = ['Phone Bill', 'Phone']
            if y == 4: #Apparel
                ct = ['Apparel', 'Clothes', 'Shoes']
            if y == 5: #Supplies
                ct = ['Supplies', 'Office', 'Bathroom']
            if y == 6: #Tech
                ct = ['Technology', 'Tech'] #...
            if y == 7: #Service
                ct = ['Service', 'Services', 'Dry Cleaning']
            if y == 8: #Health/Gym
                ct = ['Health','Gym','Medical','Med','Gnc','Vitamins', 'Health/Gym']
            if y == 9: #Haircut
                ct = ['Haircut', 'Hair', 'Fade', 'Cut', 'Lineup']
            if y == 10: #Groceries
                ct = ['Groceries','Grocery']
            if y == 11: #Other Food
                ct = ['Other Food', 'Res','Snacks','Candy','Beer','Wine','Alcohol','Liquor','Coffee','Drink']
            if y == 12: #Gas and Parking
                ct = ['Gas', 'Parking', 'Gas & Parking']
            if y == 13: #Insurance
                ct = ['Insurance', 'Ins']
            if y == 14: #Credit Card
                ct = ['Credit','Credit Card','Bank', 'Interest', 'Cc']
            if y == 15: #Student Loans
                ct = ['Student','Loans','Student Loans','Student Loan', 'FAFSA']
            if y == 16: #Vehicle Expense
                ct = ['Car','Motorcycle','Engine']
            if y == 17: #Entertainment
                ct = ['Entertainment','Movies','Sport','Sports','Game']
            if y == 18: #Subscriptions
                ct = ['Sub','Subscriptions','Xxx','Recurring','Spotify']
            if y == 19: #Travel
                ct = ['Travel','Flight','Train','Plane','Airplane','Bus']
            if y == 20: #Public Transportation
                ct = ['Public Transportation', 'Transportation', 'Uber', 'Marta', 'Cab']
            if y == 21: #Seasonal/Special
                ct = ['Special', 'Seasonal', 'Graduation', 'Spring Break', 'Recital', 'Gift', 'Gifts']


            for j in range(2,(z+1)): #Updates all active categories
                if mez['B' + str(j)].value in ct and mez['C' + str(j)].value != 'qq':
                    add = float(Sum[get_column_letter(mz+8) + str(i+2)].value) + float(mez['D' + str(j)].value)
                    Sum[get_column_letter(mz+8) + str(i+2)] = float(add)
                if g != 1 and mez['B' + str(j)].value in ct and mez['C' + str(j)].value == 'qq':
                    add = float(Sum[get_column_letter(mz+7) + str(i+2)].value) + float(mez['D' + str(j)].value)
                    Sum[get_column_letter(mz+7) + str(i+2)] = float(add)

            if mNum > 3:
 #               Sum[get_column_letter(mNum+10) + str(i+2)] = '=ROUND(AVERAGE(J'+str(i+2)+':'+get_column_letter(mNum+8)+str(i+2)+'),2)'
                Sum[get_column_letter(mNum+10)+str(i+2)] = '=ROUND(AVERAGE(J{0}:{1}{0}),2)'.format(str(i+2), get_column_letter(mNum+8))
 #               Sum[get_column_letter(mNum+11) + str(i+2)] = '=ROUND(AVERAGE('+get_column_letter(mNum+6)+str(i+2)+':'+get_column_letter(mNum+8)+str(i+2)+'),2)'
                Sum[get_column_letter(mNum+11)+str(i+2)] = '=ROUND(AVERAGE(J{0}:{1}{0}),2)'.format(str(i+2), get_column_letter(mNum+6))
            else:
 #               Sum[get_column_letter(mNum+10) + str(i+2)] = '=\'['+str(int(year)-1)+' Monthly Expenses.xlsx]Summary\'!$V$'+str(i+2)
                Sum[get_column_letter(mNum+10)+str(i+2)] = '=\'[{} Monthly Expenses.xlsx]Summary\'!$V${}'.format(str(int(year)-1), str(i+2))
 #               Sum[get_column_letter(mNum+11) + str(i+2)] = '=\'['+str(int(year)-1)+' Monthly Expenses.xlsx]Summary\'!$W$'+str(i+2)
                Sum[get_column_letter(mNum+11)+str(i+2)] = '=\'[{} Monthly Expenses.xlsx]Summary\'!$W${}'.format(str(int(year)-1), str(i+2))
                
        # Below updates the 'other' category and total for month
 #       Sum[get_column_letter(mz+8) + str(ml-1)] = '=' +TM[g]+ '!E' + str(z) + '-SUM(Summary!' +get_column_letter(mz+8)+ '2:' +get_column_letter(mz+8)+ str(ml-2)+ ')'
        Sum[get_column_letter(mz+8)+str(ml-1)] = '={0}!E{1}-SUM(Summary!{2}2:{2}{3})'.format(TM[g], str(z), get_column_letter(mz+8), str(ml-2))
 #       Sum[get_column_letter(mz+8) + str(ml+1)] = '=SUM('+get_column_letter(mz+8)+ '2:' +get_column_letter(mz+8)+ str(ml-1)+ ')'
        Sum[get_column_letter(mz+8)+str(ml+1)] = '=SUM({0}2:{0}{1})'.format(get_column_letter(mz+8), str(ml-1))
        
    # Updates the year total and average
#    Sum['I' +str(ml-1)] = '=SUM(J' +str(ml-1)+ ':'+get_column_letter(mNum+9) +str(ml-1)+ ')'
    Sum['I' +str(ml+1)] = '=SUM(J{0}:{1}{1})'.format(ml-1, get_column_letter(mNum+9))
    if mNum > 3: #Sets year average and 3-mo average when current month is after March
#        Sum[get_column_letter(mNum+10) + str(ml+1)] = '=ROUND(AVERAGE(J'+str(ml+1)+':'+get_column_letter(mNum+8)+str(ml+1)+'),2)'
        Sum[get_column_letter(mNum+10)+str(ml+1)] = '=ROUND(AVERAGE(J{0}:{1}{0}),2)'.format(ml+1, get_column_letter(mNum+8))
#        Sum[get_column_letter(mNum+11) + str(ml+1)] = '=ROUND(AVERAGE('+get_column_letter(mNum+6)+str(ml+1)+':'+get_column_letter(mNum+8)+str(ml+1)+'),2)'
        Sum[get_column_letter(mNum+11)+str(ml+1)] = '=ROUND(AVERAGE({0}{1}:{2}{1}),2)'.format(get_column_letter(mNum+6), ml+1, get_column_letter(mNum+8))
    else: #Sets year average and 3-mo average using last year's averages when month is Jan, Feb, or Mar (personal)
        Sum[get_column_letter(mNum+10) + str(ml+1)] = '=\'/Users/alexrogers823/Documents/Python Projects/['+str(int(year)-1)+' Monthly Expenses.xlsx]Summary\'!$V$'+str(ml+1)
        Sum[get_column_letter(mNum+11) + str(ml+1)] = '=\'/Users/alexrogers823/Documents/Python Projects/['+str(int(year)-1)+' Monthly Expenses.xlsx]Summary\'!$W$'+str(ml+1)

    for r in range(len(categories)): #Compares month goals to actual month number
        if Sum[get_column_letter(mNum+12)+str(r+2)].value != None and Sum[get_column_letter(mNum+9)+str(r+2)].value > Sum[get_column_letter(mNum+12)+str(r+2)].value:
            Sum[get_column_letter(mNum+12)+str(r+2)].style = Style(font=Font(name='Calibri', color='FF0000'))

# Gives cumulative price on each month tab     
def cumulativePrice():
    m = x-1
    if month['C'+str(x)].value == 'qq': #Excludes late previous-month expenses
        if x == 2:
            month['E2'] = 'Not Included'
            return
        if x > 2:
            month['E'+str(x)] = '=E'+str(x-1)
            return
    else:
        if x == 2:
            month['E2'] = '=D2'
            return
        elif x > 2 and month['E'+str(x-1)].value == 'Not Included':
            match = False
            while match == False: #Skips over any cells that say 'Not Included'
                if month['E'+str(m)].value == 'Not Included':
                    if m == 1:
                        month['E'+str(x)] = '=D'+str(x)
                    else:
                        m -= 1
                else:
                    if month['E'+str(m)].value == 'Cumulative Price':
                        month['E'+str(x)] = '=D'+str(x)
                    else:
                        month['E'+str(x)] = '=E'+str(m)
                        match == True
                    break
            return
        else:
 #           month['E' + str(x)] = '=D' + str(x) + '+E' + str((x) - 1)
            month['E'+str(x)] = '=D{}+E{}'.format(str(x), str(x-1))
            return

# Makes pie chart for each month
def monthChartBreakdown():
    curMonth = ['','January','February','March','April','May','June','July','August','September','October','November','December']
    for i in range(1, mNum+1):
        pie = PieChart()
        m = len(categories)
        monthData = Reference(Sum, min_col = i+9, min_row = 2, max_row = m-1)
        labels = Reference(Sum, min_col=8, min_row=2, max_row = m-1)
        pie.add_data(monthData)
        pie.set_categories(labels)
        pie.title = curMonth[i]+ ' Expenses by Category'
        pie.width = 18.0
        pie.height = 12.0
        pie.legend.layout = Layout(manualLayout=ManualLayout(x=0.25, y=0.25, h=0.99, w=0.25))

        wb.get_sheet_by_name(TM[i]).add_chart(pie, 'G3')

# Makes pie chart and bar chart for summary tab
def chartBreakdown():
    pie = PieChart()
    z = len(categories)

    data = Reference(Sum, min_col=9, min_row=2, max_row= z-1)
    labels = Reference(Sum, min_col=8, min_row=2, max_row= z-1)
    pie.add_data(data)
    pie.set_categories(labels)
    pie.title = 'Breakdown of Expenses'
    pie.width = 15.0
    pie.height = 12.0
    pie.legend.layout = Layout(manualLayout=ManualLayout(x=0.25, y=0.25, h=0.99, w=0.25))
    
    Sum.add_chart(pie, 'A1')
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True

    bar = BarChart()
    barData1 = Reference(Sum, min_col=mNum+9, min_row=1, max_row=z-1)
    barData2 = Reference(Sum, min_col=mNum+12, min_row=1, max_row=z-1)
    bar.add_data(barData1, titles_from_data=True)
    bar.add_data(barData2, titles_from_data=True)
    bar.set_categories(labels)
    bar.title = 'Goal Comparison'
    bar.width = 2.0*z
    bar.height = 12.0
    bar.legend.layout = Layout(manualLayout=ManualLayout(x=0.25, y=0.25, h=1.99, w=0.25))
    Sum.add_chart(bar, 'A28')

#def save_and_update():
#    wbSave = saving.Update(xl, 'open')
#    wbSave.openXL()

# Saves and opens excel workbook with new expenses added
def save_and_update():
    wb.save(curYear + '.xlsx')
    print('Would you like to see the additions to your expenses? [Y or N]')
    openXL = input()
    if openXL.lower().startswith('y'):
        opener = 'open'
        print('Opening...')
        time.sleep(2)
        subprocess.call([opener, (curYear + '.xlsx')])
  

def orderOfStatements():
    addExpense()
    cumulativePrice()
    categoryTotal()
    moreReceipts()

# Lets user set a month goal for an expense
def goals():
    k = True
    print('Enter goal for category, followed by number (Haircut, 30)')
    print()
    for i in range(len(categories)):
        print(categories[i], end=', ')
    print()
    while k == True:
        goal = input().title().split(', ')
        for c in range(len(categories)-3):
            if categories[c].startswith(goal[0]):
                Sum[get_column_letter(mNum+12)+str(c+2)] = float(goal[1])
        print('Another one?')
        if input().lower().startswith('n'):
            k = False
            break

    print('continue with Monthly Expenses?')
    return input().lower()
            
# Puts late previous-month expenses in previous month tab
def qqAdd():
    for i in range(2, mNum+1):
        search = wb.get_sheet_by_name(TM[i])
        add = wb.get_sheet_by_name(TM[i-1])
        sx = search.max_row
        ax = add.max_row
        qqTotal = float(0)
        for j in range(1, sx+1):
            if search['C'+str(j)].value == 'qq':
                qqTotal += float(search['D'+str(j)].value)
        add['E'+str(ax+2)] = float(add['E'+str(ax)].value) + qqTotal

def qqFill():
    for i in range(1, mNum+1):
        blue = wb.get_sheet_by_name(TM[i])
        for j in range(2, x+1):
            if blue['C'+str(j)].value == 'qq':
                blue['D'+str(j)].fill = PatternFill(start_color='7BBDD3', fill_type='solid')
                blue['E'+str(j)].fill = PatternFill(start_color='7BBDD3', fill_type='solid')


def moreReceipts():
    global receipts
    print('Do you have other receipts? [Y or N]')
    print()
    if input().lower().startswith('n'):
        receipts = False
        qqFill()
        monthChartBreakdown()
        chartBreakdown()
        save_and_update()
        return
    else:
        orderOfStatements()

# For small corrections when user doesn't need the entire code to be run
def shortExit():
    categoryTotal()
    monthChartBreakdown()
    chartBreakdown()
    save_and_update()

                           
#Main (Global Variables)
year = datetime.datetime.today().strftime('%Y')
mNum = int(datetime.datetime.today().strftime('%m')) #current month as a number
dNum = int(datetime.datetime.today().strftime('%d')) #current day as a number
hNum = int(datetime.datetime.today().strftime('%H')) #current hour as a number
first = False

#TM represents current month for formatting in summary tab. Jan = 1
TM = ['Year Total','Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sept','Oct','Nov','Dec']

curYear = 'GITHub {} Monthly Expenses'.format(year)
xl = curYear + '.xlsx'
try:
    wb = px.load_workbook(xl)
except FileNotFoundError:
    ny = px.load_workbook('blank.xlsx') #Have to find a way to do this out of thin air (for other users)
    ny.create_sheet(index = 0, title = 'Summary')
    ny.get_sheet_by_name('Summary')['J1'] = 'Average'
    ny.create_sheet(index = 1, title = 'Category Words')
    ny.save(xl)
    wb = px.load_workbook(xl)
Sum = wb.get_sheet_by_name('Summary')
cword = wb.get_sheet_by_name('Category Words')
month = setWorkbook()
x = month.max_row # x represents the number on the row that you're currently at
GA = formatSheet(month)
categories = setSummarySheet()
setMonth()
receipts = False
keyword = ''
if hNum < 12:
    hour = 'Morning'
elif hNum >= 12 and hNum < 17:
    hour = 'Afternoon'
else:
    hour = 'Evening'


#This is how the program starts
print('Good {}. Do you have expenses to add? [Y or N]'.format(hour))
addr = input()
if addr.lower().startswith('goal'):
    changeGoals = goals()
if addr.lower().startswith('correct'): # or ideal.startswith('n'):
    shortExit()
elif addr.lower().startswith('y'): #or ideal.startswith('y'):
    receipts == True
    print()
    reviewMaterials()
    revertStyle()
    orderOfStatements()
else:
    receipts == False




