import openpyxl as px
import time, random, datetime, webbrowser

def setFormat(): #Formatting in excel
    titlesI = ['Workout', 'Primary Muscle', 'Secondary Muscle', 'Std. Sets', 'Std. Reps', 'Start Weight', 'End Weight', 'Max Weight']
    for i in range(len(titlesI)):
        ind[chr(65+i)+'1'] = titlesI[i]
    ind.column_dimensions['A'].width = 30
    ind.column_dimensions['B'].width = 20
    ind.column_dimensions['C'].width = 20

    titlesC = ['Circuit', 'Sets', 'Muscle Target 1', 'Muscle Target 2', 'Workout Days', 'Rest Days', 'Reference']
    for i in range(len(titlesC)):
        cir[chr(65+i)+'1'] = titlesC[i]

def mainMenu():
    global mode
    mode = ''
    print('What would you like to do?')
    print('Add workout')
    print('Make this week\'s workout')
    print('Edit workout stats')
    print('Quit program')
    choice = input()
    if choice.lower().startswith('ad'):
        pre_Workout()
        addWorkout()
        return
    if choice.lower().startswith('mak'):
        pre_Workout()
        if mode == 'single':
            generate()
        if mode == 'circuit':
            circuitGenerate()
        return
    if choice.lower().startswith('ed'):
        logWeight()
        return
    if choice.lower().startswith('q'):
        return


def pre_Workout(): #selects what to generate based on user's output
    global mode, x
    while mode != 'single' or mode != 'circuit':
        try:
            print('Is this a single workout or a circuit?')
            m = input()
            if not m.lower().startswith('sin') or not m.lower().startswith('cir'):
                raise Exception
        except Exception:
            print('Must specify a single or circuit workout')
            time.sleep(1)
        else:
            if m.lower().startswith('sin'):
                mode = 'single'
                break
            elif m.lower().startswith('cir'):
                mode = 'circuit'
                x = cir.max_row
                break
    return

def addWorkout(): #Adds newly input workout to excel
    global x
    if mode == 'single':
        x += 1
        print('What\'s the workout?')
        name = input()
        ind['A'+str(x)] = name.title()
    if mode == 'circuit':
        x += 2
        print('What\'s the name of the circuit?')
        name = input()
        cir['A'+str(x)] = 'CIR {}'.format(name.title())
    print('What are the muscles it targets? (max 2)')
    mus = input().title().split()
    for i in range(len(mus)):
        if mode == 'single':
            if mus[i] == 'Back':
                print('Is it [u]pper, [l]ower, or [a]ll of your back?')
                back = input()
                if back.lower().startswith('u'):
                    ind[chr(66+i)+str(x)] = 'Upper Back'
                elif back.lower().startswith('l'):
                    ind[chr(66+i)+str(x)] = 'Lower Back'
                elif back.lower().startswith('a'):
                    ind[chr(66+i)+str(x)] = 'Back'
                else:
                    ind[chr(66+i)+str(x)] = mus[i]
            else:
                ind[chr(66+i)+str(x)] = mus[i]
        if mode == 'circuit':
            cir[chr(67+i)+str(x)] = mus[i]
    if mode == 'circuit':
        print('Number of workout days and rest days (if specified)')
        rest = input()
        rest.split()
        for r in range(len(rest)):
            cir[chr(69+i)+str(x)] = rest[r]
        print('What is the reference?')
        ref = input()
        cir['G'+str(x)] = ref
        circuitDetails()
    moreWorkouts()

def circuitDetails(): #Adds circuit to excel
    global x
    d = 1
    x += 1
    cir['A'+str(x)] = 'Day {}:'.format(str(d))
    while True:
        x += 1
        print('Enter workout in the circuit')
        sub = input()
        cir['A'+str(x)] = ' - {}'.format(sub.title())
        print('How many sets?')
        sets = input()
        cir['B'+str(x)] = sets
        print('[A]nother one, go to [n]ext day, or [e]nd?')
        yep = input()
        if yep.lower().startswith('a'):
            continue
        if yep.lower().startswith('n'):
            d += 1
            x += 2
            cir['A'+str(x)] = 'Day {}:'.format(str(d))
            continue
        if yep.lower().startswith('e'):
            break
    return


def moreWorkouts():
    print('Are there any more %s workouts?' % (mode))
    print()
    if input().lower().startswith('y'):
        addWorkout()
    else:
        mainMenu()

def generate(): #Generates random workout based on days and type of workout
    global txtlist
    typ = setWktType()
    if typ == 'dropset':
        drop = 0.9
        goal = 1.1**3
        start = 1
    elif typ == 'reignition':
        goal = 1.1**4
        start = 0.9
    else:
        goal = 1.1**4
        start = 1
    y = 0
    print('How many exercises per day?')
    exer = int(input())
    print('Generating...')
    print()
    time.sleep(3)
    print('Mode: %s' % (typ.title()))
    print()
    for i in range(1, days+1):
        txtlist.append('Day %s:' % (i))
        print('Day %s:' % (i))
        if y % 2 == 0:
            day = upper
        else:
            day = lower
        for j in range(exer):
            w = random.randint(0, len(day)-1)
            for s in range(2, x+1):
                if day[w] == ind['A'+str(s)].value and ind['F'+str(s)].value != None:
                    st = round(int(ind['F'+str(s)].value)*start)
                    # begin = ', Start Wt: '+str(st)
                    goal = round(goal*st)
                    ideal = ', Goal: '+str(goal)
                    break
                else:
                    # begin = ''
                    ideal = ''
                if day[w] == ind['A'+str(s)].value and ind['G'+str(s)].value != None:
                    end = str(ind['G'+str(s)].value)
                    break
                else:
                    # end = ''
            # print(day[w]+'%s%s%s' % (begin, end, ideal))
            print('{}, Start Wt: {}, End Wt: {}{}'.format(day[w], st, end, ideal))
            # txt = day[w]+begin+end+ideal
            txt = '{}, Start Wt: {}, End Wt: {}{}'.format(day[w], st, end, ideal)
            if typ == 'dropset' and ind['F'+str(s)].value != None:
                theDrop = str('(Drop to %s for last set)' % (round(goal*drop)))
            else:
                theDrop = ''
            txtlist.append(txt)
            if typ == 'dropset':
                txtlist.append(theDrop)
            day.remove(day[w])
        txtlist.append('')
        print()
        time.sleep(4)
        y += 1
        # if y == 2:
        #     y -= 2
    print('Do you want to search any of these workouts?')
    # lookup = input()
    if input().lower().startswith('y'):
        howTo()
    acceptWorkout()
    return
        # if y == 1:
        #     day = lower
        #     for j in range(exer):
        #         w = random.randint(0, len(lower)-1)
        #         for s in range(2, x+1):
        #             if day[w] == ind['A'+str(s)].value and ind['F'+str(s)].value != None:
        #                 st = round(int(ind['F'+str(s)].value)*start)
        #                 begin = ', Start Wt: '+str(st)
        #                 goal *= st
        #                 goal = round(goal)
        #                 ideal = ', Goal: '+str(goal)
        #                 break
        #             else:
        #                 begin = ''
        #                 ideal = ''
        #             if day[w] == ind['A'+str(s)].value and ind['G'+str(s)].value != None:
        #                 end = ', End Wt: '+str(ind['G'+str(s)].value)
        #                 break
        #             else:
        #                 end = ''
        #         print(day[w]+'%s%s%s' % (begin, end, ideal))
        #         downTxt = day[w]+begin+end+ideal
        #         if typ == 'dropset' and ind['F'+str(s)].value != None:
        #             theDrop = str('(Drop to %s for last set)' % (round(goal*drop)))
        #         else:
        #             theDrop = ''
        #         txtlist.append(downTxt)
        #         if typ == 'dropset':
        #             txtlist.append(theDrop)
        #         day.remove(day[w])

def circuitGenerate(): #Generates circuit workout
    global txtlist
    print('Which circuit are you trying to do?')
    print()
    for i in range(2, x+1):
        if str(cir['A'+str(i)].value).startswith('CIR'):
            print(i, cir['A'+str(i)].value.strip('CIR '), cir['G'+str(i)].value, '('+cir['C'+str(i)].value+')')
    ans = int(input())
    txtlist.append(cir['A'+str(ans)].value)
    ans += 1
    print('Generating...')
    print()
    time.sleep(3)
    for c in range(ans, x):
        if cir['A'+str(c)].value is None:
            print()
            continue #why do we need this?
        elif not cir['A'+str(c)].value.startswith('CIR'):
            print(cir['A'+str(c)].value, cir['B'+str(c)].value)
            if str(cir['A'+str(c)].value).startswith('Day'):
                txtlist.append(cir['A'+str(c)].value)
            else:
                # cycle = cir['A'+str(c)].value+', Sets: '+cir['B'+str(c)].value
                cycle = '{}, Sets: {}'.format(cir['A'+str(c)].value, cir['B'+str(c)].value)
                txtlist.append(cycle)
        else:
            break
    acceptWorkout()

def acceptWorkout():
    print('Do you accept this workout?')
    # go = input()
    if input().lower().startswith('n'):
        if mode == 'single':
            generate()
        if mode == 'circuit':
            circuitGenerate()
    else:
        txtfile()
        return

def howTo(): #Looks up workouts on YouTube
    print('Which ones? (Pick the corresponding number)')
    print()
    for i in range(len(txtlist)):
        print(i, txtlist[i])

    searches = input().split()

    for k in range(len(searches)):
        yNum = int(searches[k])
        yTube = txtlist[yNum]
        webbrowser.open('https://www.youtube.com/results?search_query=%s' % (yTube.replace(' ','+')))
    return


def setWktType(typ='standard'):
    print('Your workout is set to standard. Is this okay?')
    if input().lower().startswith('n'):
        print('Which mode do you want?')
        for i in range(len(types)):
            print(types[i])
        decide = input()
        if decide.lower().startswith('s'): #Standard or 6-10
            typ = 'standard'
        elif decide.lower().startswith('d'): #Dropset
            typ = 'dropset'
        elif decide.lower().startswith('p'): #Pyramid
            typ = 'pyramid'
        elif decide.lower().startswith('rei'): #Reignition
            typ = 'reignition'
        else:
            typ = 'standard'
    return typ


def txtfile(): #Prints generated workout to a text file
    if mode == 'single':
        sun = (datetime.timedelta(days=6)).days - datetime.datetime.today().weekday()
        top = datetime.datetime.now() + datetime.timedelta(days=sun)
        btm = datetime.datetime.now() + datetime.timedelta(days=sun+6)
        nxtSun = top.strftime('%b %d')
        nxtSat = btm.strftime('%b %d')
        with open(nxtSun+'-'+nxtSat+'.txt', 'wt') as f:
            for i in range(len(txtlist)):
                print(txtlist[i], file=f)
    if mode == 'circuit':
        cir = txtlist[0]
        with open(cir+'.txt', 'wt') as f:
            for i in range(len(txtlist)):
                print(txtlist[i], file=f)


def logWeight():
    display = []
    print('What muscle does this workout target?')
    group = input()
    for i in range(2, x+1):
        if ind['B'+str(i)].value == group.title():
            display.append(ind['A'+str(i)].value)
    print('Which workout are you putting in stats for?')
    print()
    for i in range(len(display)):
        print(str(i+1)+'. '+display[i])
    wk = int(input())-1
    for i in range(2, x+1):
        if ind['A'+str(i)].value == display[wk]:
            d = i
    print('Which weight? [Start Wt, End Wt, Max Wt]')
    wt = input()
    print('What number?')
    num = int(input())
    if wt.lower().startswith('s'):
        ind['F'+str(d)] = num
    elif wt.lower().startswith('e'):
        ind['G'+str(d)] = num
    elif wt.lower().startswith('m'):
        ind['H'+str(d)] = num
    else:
        print('Must specify which weight. Try again') #convert this into a try/else
        time.sleep(1)
        mainMenu()

def upperWkts():
    upperList = []
    for i in range(2, x+1):
        if ind['B'+str(i)].value in armList or ind['B'+str(i)].value in shoulderList or ind['B'+str(i)].value in chestList or ind['B'+str(i)].value in ubackList:
            upperList.append(ind['A'+str(i)].value)
    return upperList

def lowerWkts():
    lowerList = []
    for i in range(2, x+1):
        if ind['B'+str(i)].value in lbackList or ind['B'+str(i)].value in buttList or ind['B'+str(i)].value in thighList or ind['B'+str(i)].value in legList:
            lowerList.append(ind['A'+str(i)].value)
    return lowerList

def allMuscles():
    mas = []
    for i in range(len(armList)):
        mas.append(armList[i])
    for i in range(len(shoulderList)):
        mas.append(shoulderList[i])
    for i in range(len(chestList)):
        mas.append(chestList[i])
    for i in range(len(ubackList)):
        mas.append(ubackList[i])
    for i in range(len(lbackList)):
        mas.append(lbackList[i])
    for i in range(len(buttList)):
        mas.append(buttList[i])
    for i in range(len(thighList)):
        mas.append(thighList[i])
    for i in range(len(legList)):
        mas.append(legList[i])
    return mas


#Main
#wb = px.load_workbook('List of Workouts TEST.xlsx')
wb = px.load_workbook('List of Workouts.xlsx')
ind = wb.get_sheet_by_name('Individual Workouts')
cir = wb.get_sheet_by_name('Circuit Workouts')

#masterList = allMuscles()
masterList = {
    'Arms': ['Arms', 'Biceps', 'Triceps', 'Forearms'],
    'Shoulders': ['Shoulders', 'Delts', 'Traps'],
    'Chest': ['Chest', 'Pecs'],
    'UpperBack': ['Back', 'Upper Back', 'Lats'],
    'LowerBack': ['Back', 'Lower Back'],
    'Butt': ['Butt', 'Gluts'],
    'Thighs': ['Thighs', 'Quads', 'Hamstrings'],
    'Legs': ['Legs', 'Calf', 'Calves']
    }

setFormat()
x = ind.max_row
days = 4

upper = upperWkts()
lower = lowerWkts()
#bench = benchPress()
#sc = specialInstructions()
mode = ''
txtlist = []
types = ['Standard (6-10)','Pyramid','Dropset','Reignition']

mainMenu()
wb.save('List of Workouts.xlsx')
